"""
spreadsheet_db.py

Small read-only adapter for the MDB 2026-27 Excel workbook.
It gives the router/data layers a database-like view of student rows before
falling back to Moodle SQL.
"""

from __future__ import annotations

import os
import re
from dataclasses import dataclass
from datetime import date, datetime
from functools import lru_cache
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_WORKBOOK_PATH = BASE_DIR / "data" / "MDB_2026_27_Formate.xlsx"
WORKBOOK_PATH = Path(os.getenv("ACADEMIC_DB_XLSX_PATH", DEFAULT_WORKBOOK_PATH))
PRIMARY_SHEET = os.getenv("ACADEMIC_DB_XLSX_SHEET", "MDB 2026-27")

USN_RE = re.compile(r"\b[0-9][a-z]{2}[0-9]{2}[a-z]{2}[0-9]{3}\b", re.IGNORECASE)


HEADER_ALIASES = {
    "Sl.No": "serial_no",
    "Name": "name",
    "USN": "usn",
    "M/F": "gender",
    "DOB": "dob",
    "X %(only mention in %)": "x_percent",
    "XII %(only mention in %)": "xii_percent",
    "No. of Year Gaps after 12th or Diploma": "year_gaps",
    "No. of Current Backlogs in B.E(mentione backlogs in numbers only)": "backlog_count",
    "Current Backlogs in B.E(mentione backlogs subject name only)": "backlog_subjects",
    "I (SGPA)": "sgpa_1",
    "II (SGPA)": "sgpa_2",
    "III (SGPA)": "sgpa_3",
    "IV (SGPA)": "sgpa_4",
    "V (SGPA)": "sgpa_5",
    "CGPA(Till 5th sem)": "cgpa",
    "Agg.%(CGPA)*10": "aggregate_percent",
    "Mobile No": "mobile",
    "E-mail Id(College Email Address)": "college_email",
    "Email ID (Personal Email Address)": "personal_email",
    "Area of Interest": "interests",
}

GRADE_KEYWORDS = {
    "cgpa", "sgpa", "grade", "grades", "marks", "score", "percentage", "percent",
    "aggregate", "xii", "12th", "tenth", "10th",
}
CONTACT_KEYWORDS = {"contact", "phone", "mobile", "email", "whatsapp"}
BACKLOG_KEYWORDS = {"backlog", "backlogs", "arrear", "arrears"}
PROFILE_KEYWORDS = {
    "profile", "details", "student", "name", "usn", "dob", "interest",
    "interests", "gap", "gender",
}


@dataclass(frozen=True)
class SpreadsheetMatch:
    intent: Optional[str]
    entity: Optional[str]
    student: Optional[Dict[str, Any]]
    matched_by: str
    confidence: float


def _clean_header(value: Any) -> str:
    text = str(value or "").replace("\n", " ").strip()
    return re.sub(r"\s+", " ", text)


def _clean_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str):
        text = value.strip()
        return text if text else None
    return value


def _normalise_usn(value: Any) -> str:
    return str(value or "").strip().upper()


def _normalise_name(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def _workbook_signature() -> tuple[str, float]:
    path = WORKBOOK_PATH
    if not path.exists():
        return (str(path), 0)
    return (str(path), path.stat().st_mtime)


@lru_cache(maxsize=4)
def _load_records_cached(path_text: str, mtime: float) -> List[Dict[str, Any]]:
    path = Path(path_text)
    if not path.exists():
        return []

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[PRIMARY_SHEET] if PRIMARY_SHEET in workbook.sheetnames else workbook.worksheets[0]

    raw_headers = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
    columns: Dict[int, str] = {}
    for idx, raw in enumerate(raw_headers):
        cleaned = _clean_header(raw)
        key = HEADER_ALIASES.get(cleaned)
        if key:
            columns[idx] = key

    records: List[Dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=3, values_only=True):
        record = {
            field: _clean_value(row[idx]) if idx < len(row) else None
            for idx, field in columns.items()
        }
        if not record.get("usn") and not record.get("name"):
            continue
        record["usn"] = _normalise_usn(record.get("usn"))
        records.append(record)

    workbook.close()
    return records


def load_records() -> List[Dict[str, Any]]:
    """Load workbook rows with an mtime-aware cache."""
    return _load_records_cached(*_workbook_signature())


def student_count() -> int:
    return len(load_records())


def _find_student_by_usn(usn: str) -> Optional[Dict[str, Any]]:
    target = _normalise_usn(usn)
    for record in load_records():
        if record.get("usn") == target:
            return record
    return None


def _find_student_by_name(query: str) -> Optional[Dict[str, Any]]:
    lowered = _normalise_name(query)
    if not lowered:
        return None

    best: tuple[int, Optional[Dict[str, Any]]] = (0, None)
    query_tokens = set(re.findall(r"[a-z]+", lowered))
    for record in load_records():
        name = _normalise_name(record.get("name"))
        if not name:
            continue
        if name in lowered:
            return record
        name_tokens = set(re.findall(r"[a-z]+", name))
        score = len(query_tokens & name_tokens)
        if score > best[0] and score >= 2:
            best = (score, record)
    return best[1]


def find_student(entity: str) -> Optional[Dict[str, Any]]:
    if not entity or entity.lower() == "general":
        return None
    usn_match = USN_RE.search(entity)
    if usn_match:
        return _find_student_by_usn(usn_match.group(0))
    return _find_student_by_usn(entity) or _find_student_by_name(entity)


def _intent_from_query(query: str) -> Optional[str]:
    lowered = query.lower()
    tokens = set(re.findall(r"[a-z0-9]+", lowered))
    if tokens & BACKLOG_KEYWORDS:
        return "backlog_report"
    if tokens & CONTACT_KEYWORDS:
        return "contact_lookup"
    if tokens & GRADE_KEYWORDS:
        return "student_profile"
    if tokens & PROFILE_KEYWORDS:
        return "student_profile"
    return None


def inspect_query(query: str) -> SpreadsheetMatch:
    """
    Read the workbook before data retrieval and infer whether it can sharpen
    the classifier output.
    """
    usn_match = USN_RE.search(query)
    if usn_match:
        student = _find_student_by_usn(usn_match.group(0))
        if student:
            return SpreadsheetMatch(
                intent=_intent_from_query(query) or "student_profile",
                entity=student["usn"],
                student=student,
                matched_by="usn",
                confidence=0.95,
            )

    student = _find_student_by_name(query)
    if student:
        return SpreadsheetMatch(
            intent=_intent_from_query(query) or "student_profile",
            entity=student["usn"],
            student=student,
            matched_by="name",
            confidence=0.75,
        )

    return SpreadsheetMatch(
        intent=_intent_from_query(query),
        entity=None,
        student=None,
        matched_by="keyword",
        confidence=0.4 if _intent_from_query(query) else 0.0,
    )


def enrich_classification(query: str, classification: Dict[str, str]) -> Dict[str, str]:
    """
    Modify router classification using facts read from the Excel database.
    This is intentionally conservative: workbook matches only override generic
    or weak classifications, or replace entity with a known USN.
    """
    enriched = dict(classification)
    match = inspect_query(query)

    if match.student:
        enriched["query_type"] = "organizational_query"
        enriched["entity"] = match.entity or enriched.get("entity", "general")
        if enriched.get("intent") in {"general", "", None} or match.intent:
            enriched["intent"] = match.intent or "student_profile"
        enriched["source"] = "xlsx_db"
        enriched["matched_by"] = match.matched_by
    elif match.intent and enriched.get("intent") in {"general", "", None}:
        enriched["query_type"] = "organizational_query"
        enriched["intent"] = match.intent
        enriched["source"] = "xlsx_db"

    return enriched


def student_summary(student: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "student_id": student.get("usn"),
        "name": student.get("name"),
        "username": student.get("usn"),
        "email": student.get("college_email"),
        "phone": student.get("mobile"),
        "gender": student.get("gender"),
        "dob": student.get("dob"),
        "x_percent": student.get("x_percent"),
        "xii_percent": student.get("xii_percent"),
        "year_gaps": student.get("year_gaps"),
        "backlog_count": student.get("backlog_count") or 0,
        "backlog_subjects": student.get("backlog_subjects"),
        "sgpa": {
            "I": student.get("sgpa_1"),
            "II": student.get("sgpa_2"),
            "III": student.get("sgpa_3"),
            "IV": student.get("sgpa_4"),
            "V": student.get("sgpa_5"),
        },
        "cgpa": student.get("cgpa"),
        "aggregate_percent": student.get("aggregate_percent"),
        "interests": student.get("interests"),
        "source": "MDB 2026-27 Excel database",
    }


def user_context_for_student(student: Dict[str, Any]) -> Dict[str, Any]:
    profile = {
        "id": student.get("usn"),
        "student_id": student.get("usn"),
        "name": student.get("name"),
        "username": student.get("usn"),
        "email": student.get("college_email"),
        "phone": student.get("mobile"),
        "department": "ISE",
        "semester": 5,
        "section": "—",
        "course": "MDB 2026-27",
        "cgpa": student.get("cgpa"),
        "attendance_percent": "—",
        "backlog_count": student.get("backlog_count") or 0,
        "mentor": {"name": "Not assigned", "email": "", "phone": ""},
        "class_teacher": {"name": "Not available", "email": "", "phone": ""},
    }
    return {
        "role": "student",
        "user_id": student.get("usn"),
        "profile": profile,
        "overview": {
            "students": student_count(),
            "courses": ["MDB 2026-27"],
        },
        "permissions": {
            "can_assign_mentor": False,
            "can_view_all_students": False,
        },
        "source": "MDB 2026-27 Excel database",
    }


def contact_summary(student: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "student_id": student.get("usn"),
        "name": student.get("name"),
        "student_contact": {
            "email": student.get("college_email"),
            "personal_email": student.get("personal_email"),
            "phone": student.get("mobile"),
        },
        "source": "MDB 2026-27 Excel database",
    }


def backlog_summary(student: Dict[str, Any]) -> Dict[str, Any]:
    count = student.get("backlog_count") or 0
    subjects = student.get("backlog_subjects")
    if isinstance(subjects, str):
        courses = [part.strip() for part in re.split(r"[,/]", subjects) if part.strip()]
    else:
        courses = []
    return {
        "count_with_backlogs": 1 if count else 0,
        "students": [{
            "student_id": student.get("usn"),
            "name": student.get("name"),
            "backlog_count": count,
            "backlog_courses": courses,
        }] if count else [],
        "source": "MDB 2026-27 Excel database",
    }
